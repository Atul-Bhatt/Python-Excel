from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# An excel document can have one or more worksheets
wb = Workbook()
ws = wb.active

# Create a new worksheet
ws1 = wb.create_sheet('SecondSheet')
ws2 = wb.create_sheet('FirstSheet', 0)

ws.title = 'CurrentSheet'

# wb.sheetnames returns unicode strings so we have to convert them into ascii strings
# print([x.encode('ascii') for x in wb.sheetnames])

# Open an existing workbook

wb2 = load_workbook('regions.xlsx')
new_sheet = wb2.create_sheet('NewSheet')

active_sheet = wb2.active
# cell = active_sheet['A1']
# print(cell.value)
active_sheet['A1'] = 0
wb2.save('NewRegions.xlsx')
